import json
import logging
from django.http import JsonResponse, HttpResponseForbidden, HttpResponseBadRequest
from django.views.decorators.csrf import csrf_exempt
from django.conf import settings
from .models import SupStorico, ColorServices
from django.views.generic import ListView
from django.db.models import Q, Sum
from django.http import HttpResponse
from django.template.loader import get_template
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from weasyprint import HTML, CSS
from weasyprint.text.fonts import FontConfiguration
import io
from datetime import datetime

logger = logging.getLogger('django')


@csrf_exempt
def get_data_color_service(request):
    if request.method != 'POST':
        return HttpResponseBadRequest("Only POST method allowed")

    expected_token = getattr(settings, 'API_TOKEN', 'color_service_supersecret')
    auth_header = request.headers.get('Authorization')

    if not auth_header or auth_header != f"Bearer {expected_token}":
        return HttpResponseForbidden("Invalid or missing token")

    try:
        data = json.loads(request.body.decode('utf-8'))
    except json.JSONDecodeError:
        return HttpResponseBadRequest("Invalid JSON")

    logger.info(f"Received data: {json.dumps(data)[:500]}")

    # Очистка таблицы
    SupStorico.objects.all().delete()

    # Создание экземпляров модели
    objects_to_create = []
    for item in data:
        obj = SupStorico(
            sup_storico_id=item.get("SUP_Storico_id"),
            macchina=item.get("Macchina"),
            lotto=item.get("Lotto"),
            introduzione=item.get("Introduzione"),
            tank=item.get("Tank"),
            data_dosaggio=item.get("Data_Dosaggio"),
            data_inizio=item.get("Data_Inizio"),
            data_fine=item.get("Data_Fine"),
            id_prodotto=item.get("ID_Prodotto"),
            da_dosare=item.get("DaDosare"),
            dosato=item.get("Dosato"),
            macchina_cs=item.get("MacchinaCS"),
            linea=item.get("Linea"),
            tipo_chiamata=item.get("TipoChiamata"),
            so_number=item.get("SOnumber"),
            po_number=item.get("POnumber")
        )
        objects_to_create.append(obj)

    SupStorico.objects.bulk_create(objects_to_create, batch_size=1000)

    return JsonResponse({
        "status": "ok",
        "cleared": True,
        "inserted_records": len(objects_to_create)
    })


class SupStoricoListView(ListView):
    model = SupStorico
    template_name = 'sup_storico/sup_storico_list.html'
    context_object_name = 'records'
    paginate_by = 25

    def get_queryset(self):
        queryset = SupStorico.objects.all()
        search_query = self.request.GET.get('search')

        if search_query:
            # Поиск по связанным продуктам ColorServices
            color_services_ids = ColorServices.objects.filter(
                Q(title__icontains=search_query) |
                Q(type__icontains=search_query)
            ).values_list('product_name', flat=True)

            queryset = queryset.filter(
                Q(macchina__icontains=search_query) |
                Q(lotto__icontains=search_query) |
                Q(id_prodotto__icontains=search_query) |
                Q(tank__icontains=search_query) |
                Q(so_number__icontains=search_query) |
                Q(po_number__icontains=search_query) |
                Q(id_prodotto__in=color_services_ids)
            )

        macchina_filter = self.request.GET.get('macchina')
        if macchina_filter:
            queryset = queryset.filter(macchina=macchina_filter)

        date_from = self.request.GET.get('date_from')
        date_to = self.request.GET.get('date_to')

        if date_from:
            queryset = queryset.filter(data_dosaggio__date__gte=date_from)
        if date_to:
            queryset = queryset.filter(data_dosaggio__date__lte=date_to)

        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        filtered_queryset = self.get_queryset()

        # Создаем словарь ColorServices для оптимизации
        color_services_dict = {}
        for cs in ColorServices.objects.all():
            color_services_dict[cs.product_name] = {
                'title': cs.title,
                'type': cs.type
            }

        # Добавляем информацию о продуктах к каждой записи
        records_with_products = []
        for record in context['records']:
            record.color_service_info = color_services_dict.get(record.id_prodotto, None)
            records_with_products.append(record)
        context['records'] = records_with_products

        total_dosato = filtered_queryset.aggregate(
            total_dosato=Sum('dosato')
        )['total_dosato'] or 0

        total_da_dosare = filtered_queryset.aggregate(
            total_da_dosare=Sum('da_dosare')
        )['total_da_dosare'] or 0

        records_with_dosato = filtered_queryset.filter(dosato__isnull=False).count()
        context['machines'] = SupStorico.objects.values_list('macchina', flat=True).distinct().order_by('macchina')
        context['search_query'] = self.request.GET.get('search', '')
        context['selected_macchina'] = self.request.GET.get('macchina', '')
        context['date_from'] = self.request.GET.get('date_from', '')
        context['date_to'] = self.request.GET.get('date_to', '')
        context['total_dosato'] = total_dosato
        context['total_da_dosare'] = total_da_dosare
        context['records_with_dosato'] = records_with_dosato
        context['filtered_count'] = filtered_queryset.count()

        return context


class SupStoricoPDFExportView(SupStoricoListView):
    """PDF экспорт с теми же фильтрами"""

    def get(self, request, *args, **kwargs):
        queryset = self.get_queryset()

        # Получаем ColorServices данные
        color_services_dict = {}
        for cs in ColorServices.objects.all():
            color_services_dict[cs.product_name] = {
                'title': cs.title,
                'type': cs.type
            }

        # Добавляем информацию о продуктах к записям
        records_with_products = []
        for record in queryset:
            record.color_service_info = color_services_dict.get(record.id_prodotto, None)
            records_with_products.append(record)

        total_dosato = queryset.aggregate(total_dosato=Sum('dosato'))['total_dosato'] or 0
        total_da_dosare = queryset.aggregate(total_da_dosare=Sum('da_dosare'))['total_da_dosare'] or 0
        records_with_dosato = queryset.filter(dosato__isnull=False).count()

        context = {
            'records': records_with_products,
            'total_count': queryset.count(),
            'total_dosato': total_dosato,
            'total_da_dosare': total_da_dosare,
            'records_with_dosato': records_with_dosato,
            'search_query': request.GET.get('search', ''),
            'selected_macchina': request.GET.get('macchina', ''),
            'date_from': request.GET.get('date_from', ''),
            'date_to': request.GET.get('date_to', ''),
            'export_date': datetime.now(),
        }

        template = get_template('sup_storico/pdf_export.html')
        html_string = template.render(context)
        font_config = FontConfiguration()
        html = HTML(string=html_string)
        css = CSS(string='''
            @page {
                size: A4 landscape;
                margin: 1cm;
            }
            body {
                font-family: DejaVu Sans;
                font-size: 10px;
            }
            .table {
                width: 100%;
                border-collapse: collapse;
            }
            .table th, .table td {
                border: 1px solid #ddd;
                padding: 4px;
                text-align: left;
            }
            .table th {
                background-color: #f8f9fa;
                font-weight: bold;
            }
            .badge {
                padding: 2px 6px;
                border-radius: 3px;
                font-size: 9px;
            }
            .bg-primary { background-color: #007bff; color: white; }
            .bg-secondary { background-color: #6c757d; color: white; }
            .bg-success { background-color: #28a745; color: white; }
            .bg-info { background-color: #17a2b8; color: white; }
            .bg-warning { background-color: #ffc107; color: black; }
            .text-center { text-align: center; }
            .fw-bold { font-weight: bold; }
            .small { font-size: 8px; }
            .stats-box {
                background-color: #f8f9fa;
                border: 1px solid #dee2e6;
                padding: 10px;
                margin: 10px 0;
                border-radius: 5px;
            }
            .stat-item {
                display: inline-block;
                margin-right: 20px;
                padding: 5px 10px;
                background-color: white;
                border-radius: 3px;
                border: 1px solid #ddd;
            }
            .product-title {
                font-size: 8px;
                color: #666;
                margin-top: 2px;
            }
        ''', font_config=font_config)

        pdf = html.write_pdf(stylesheets=[css], font_config=font_config)
        response = HttpResponse(pdf, content_type='application/pdf')
        filename = f'color_service_export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'

        return response


class SupStoricoExcelExportView(SupStoricoListView):
    """Excel экспорт с теми же фильтрами"""

    def get(self, request, *args, **kwargs):
        queryset = self.get_queryset()

        # Получаем ColorServices данные
        color_services_dict = {}
        for cs in ColorServices.objects.all():
            color_services_dict[cs.product_name] = {
                'title': cs.title,
                'type': cs.type
            }

        total_dosato = queryset.aggregate(total_dosato=Sum('dosato'))['total_dosato'] or 0
        total_da_dosare = queryset.aggregate(total_da_dosare=Sum('da_dosare'))['total_da_dosare'] or 0
        records_with_dosato = queryset.filter(dosato__isnull=False).count()

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "История дозирования"

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        stats_font = Font(bold=True, color="FFFFFF")
        stats_fill = PatternFill(start_color="28a745", end_color="28a745", fill_type="solid")

        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Статистика в верхней части
        ws.cell(row=1, column=1, value="СТАТИСТИКА ДОЗИРОВАНИЯ").font = stats_font
        ws.cell(row=1, column=1).fill = stats_fill
        ws.merge_cells('A1:E1')
        ws.cell(row=2, column=1, value="Общее количество записей:")
        ws.cell(row=2, column=2, value=queryset.count())
        ws.cell(row=2, column=3, value="Записей с дозированием:")
        ws.cell(row=2, column=4, value=records_with_dosato)
        ws.cell(row=3, column=1, value="Общий план (да_досаре):")
        ws.cell(row=3, column=2, value=float(total_da_dosare))
        ws.cell(row=3, column=3, value="Общий факт (досато):")
        ws.cell(row=3, column=4, value=float(total_dosato))

        start_row = 5

        headers = [
            'ID', 'Машина', 'Лот', 'Продукт', 'Наименование продукта', 'Тип продукта', 'Резервуар',
            'Дата дозирования', 'Время дозирования', 'Дата начала', 'Время начала',
            'Дата окончания', 'Время окончания', 'Количество для дозирования',
            'Дозированное количество', 'Линия', 'SO номер', 'PO номер', 'Тип вызова'
        ]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=start_row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border

        for row, record in enumerate(queryset, start_row + 1):
            # Получаем информацию о продукте
            product_info = color_services_dict.get(record.id_prodotto, {})

            data = [
                record.sup_storico_id,
                record.macchina or '',
                record.lotto or '',
                record.id_prodotto or '',
                product_info.get('title', ''),  # Наименование продукта
                product_info.get('type', ''),  # Тип продукта
                record.tank or '',
                record.data_dosaggio.date() if record.data_dosaggio else '',
                record.data_dosaggio.time() if record.data_dosaggio else '',
                record.data_inizio.date() if record.data_inizio else '',
                record.data_inizio.time() if record.data_inizio else '',
                record.data_fine.date() if record.data_fine else '',
                record.data_fine.time() if record.data_fine else '',
                record.da_dosare or '',
                record.dosato or '',
                record.linea or '',
                record.so_number or '',
                record.po_number or '',
                record.tipo_chiamata or '',
            ]

            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.border = border

                if col in [1, 14, 15, 16]:  # ID, количества, линия
                    cell.alignment = Alignment(horizontal="right")
                elif col in [8, 9, 10, 11, 12, 13]:  # даты и времена
                    cell.alignment = Alignment(horizontal="center")

        # Автоширина колонок
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        # Информация об экспорте
        info_row = len(queryset) + start_row + 2
        ws.cell(row=info_row, column=1, value="Параметры экспорта:").font = Font(bold=True)

        if request.GET.get('search'):
            ws.cell(row=info_row + 1, column=1, value=f"Поиск: {request.GET.get('search')}")
        if request.GET.get('macchina'):
            ws.cell(row=info_row + 2, column=1, value=f"Машина: {request.GET.get('macchina')}")
        if request.GET.get('date_from'):
            ws.cell(row=info_row + 3, column=1, value=f"Дата с: {request.GET.get('date_from')}")
        if request.GET.get('date_to'):
            ws.cell(row=info_row + 4, column=1, value=f"Дата по: {request.GET.get('date_to')}")

        ws.cell(row=info_row + 5, column=1, value=f"Дата экспорта: {datetime.now().strftime('%d.%m.%Y %H:%M:%S')}")

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f'color_service_export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'

        return response
